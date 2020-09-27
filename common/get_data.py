# -*- coding:utf8 -*- #
# -----------------------------------------------------------------------------------
# Author: LiYangYang
# Date: 2020 年
# Description:
# -----------------------------------------------------------------------------------

import json
from configparser import ConfigParser

import openpyxl
import yaml

from common.get_path import GetPath


class GetYamlData:
    def get_yaml_data(self, filepath):
        with open(filepath, 'r', encoding='utf8') as yaml_file:
            return yaml.safe_load(yaml_file.read())  # 读取yaml文件内容,返回字典


class GetExcelData:
    """
    读取Excel数据
    """

    def get_excel_data(self, file, sheet_name):
        global line_data_tuple
        wb = openpyxl.load_workbook(file)  # 打开Excel文件
        sheet = wb[sheet_name]  # 选中sheet页
        line_data = []  # 定义空列表以添加每行数据
        all_data = []  # 定义空列表以添加全部数据
        status = True  # 定义status初始值True
        for one_row_data in sheet:
            if status:
                status = False
                continue
            for cell_data in one_row_data:
                if cell_data.value is None:
                    line_data.append('')
                else:
                    line_data.append(cell_data.value)  # 获取每个单元格数据,并添加到line_data列表
                line_data_tuple = tuple(line_data)  # 将line_data转成元组
            line_data.clear()  # 清空列表
            all_data.append(line_data_tuple)  # 添加元组数据进空元组列表
        return all_data  # 返回元组列表


class ExcelDataAlone:

    def __init__(self, file_path, sheet_name, read_only=False):
        self.__wb = openpyxl.load_workbook(file_path, read_only=read_only)  # 打开Excel
        self.__ws = self.__wb[sheet_name]  # 选中指定的sheet页
        self.path = file_path
    def get_max_row(self):
        return self.__ws.max_row

    def get_cell_data(self, column, row: int):
        return self.__ws[column + str(row)].value

    def edit_cell_data(self, column, row: int, value):
       self.__ws[column+str(row)] = value
       self.__wb.save(self.path)


class GetIniData:
    __CONFIG_PATH = r"AutoTesting\config\baseconfig.ini"

    def __init__(self, ini_path=__CONFIG_PATH):
        self.__get_path = GetPath()
        self.__sysPath = self.__get_path.get_path(ini_path)
        self.__config = ConfigParser()
        self.__config.read(self.__sysPath, encoding="utf-8")

    def get_ini_data(self, *args):
        """
        获取ini文件的内容。
        :param args:
        :return:
        """
        try:
            return self.__config.get(*args)
        except TypeError:
            raise ValueError("请传入正确的参数：(section,option)")


class GetJsonData:
    def get_json_data(self, path):
        with open(path, encoding='utf-8') as fp:
            return json.load(fp)


if __name__ == '__main__':
    # ini = GetIniData("GmpAutoTesting\\baseconfig.ini")
    # print(ini.get_ini_data())
    ex = ExcelDataAlone(r"D:\AutoTesting\gmp_rms\process_handling\data\test_to_do_items.xlsx", "test_o1_basic_information")
    element = ex.get_max_row()

    ex.edit_cell_data("A", 10, "zzzz")
    print("分割")
    print(ex.get_cell_data("A", "5"))

    # print(ex.get_cell_data('A',2))
    # element = GetJsonData().get_json_data(r'D:\Project\GmpAutoTesting\element\case_data.json')
    # case = element['rms']['document_preparation']['文件新增']
    # print(case)
    # # print(str.split())
    # print(case.split(''))
    # print(type(case))
